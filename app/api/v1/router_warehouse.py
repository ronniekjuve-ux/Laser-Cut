# -*- coding: utf-8 -*-
import json
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc, func
from datetime import datetime, timezone
from app.db.base import get_db
from app.db.models import WarehouseItem, WarehouseMovement, WarehouseRemnant, ItemNote, Application, ApplicationLayout, Customer
from app.models.user import User, UserRole
from app.core.deps import require_role
from app.schemas.warehouse import (
    WarehouseItemCreate, WarehouseItemUpdate,
    WarehouseDeductRequest, WarehouseReturnRequest,
    RemnantSplitRequest,
)

router = APIRouter(prefix="/warehouse", tags=["Warehouse"])


def _item_to_dict(item: WarehouseItem) -> dict:
    computed_size = item.size
    if not computed_size and item.sheet_w and item.sheet_h:
        computed_size = f"{int(item.sheet_w)}x{int(item.sheet_h)}"
    return {
        "id": item.id,
        "metal": item.metal,
        "grade": item.grade,
        "thickness": item.thickness,
        "size": computed_size,
        "sheet_w": item.sheet_w,
        "sheet_h": item.sheet_h,
        "sheet_count": item.sheet_count,
        "weight": item.weight,
        "min_quantity": item.min_quantity,
        "article": item.article,
        "parent_article": item.parent_article,
        "parent_sheet_w": item.parent_sheet_w,
        "parent_sheet_h": item.parent_sheet_h,
        "is_rectangular": item.is_rectangular if item.is_rectangular is not None else True,
        "vertices": item.vertices if isinstance(item.vertices, list) else None,
        "item_type": item.item_type or "standard",
        "owner": item.owner,
        "note": item.note,
        "last_deducted_at": item.last_deducted_at.isoformat() if item.last_deducted_at else None,
        "created_at": item.created_at.isoformat() if item.created_at else None,
    }


def _generate_article_base(metal: str, grade: str | None, thickness: float | None) -> str:
    """Generate article base like 'ст3/10мм'"""
    grade_part = (grade or metal or "XX").strip().lower()[:3]
    thickness_part = f"{int(thickness)}мм" if thickness else "XX"
    return f"{grade_part}/{thickness_part}"


def _generate_article(metal: str, grade: str | None, w: float | None, h: float | None, item_type: str) -> str:
    grade_part = (grade or metal or "XX").replace(" ", "").upper()[:3]
    w_part = int(w) if w else 0
    h_part = int(h) if h else 0
    suffix = "-ОСТ" if item_type == "remnant" else ""
    return f"{grade_part}-{w_part}x{h_part}{suffix}"


@router.get("/")
async def list_warehouse(
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR, UserRole.ACCOUNTANT))
):
    result = await db.execute(
        select(WarehouseItem).order_by(desc(WarehouseItem.created_at))
    )
    items = result.scalars().all()

    # Build binding map: warehouse_item_id → list of layout codes
    bindings_map = {}
    layouts_result = await db.execute(
        select(ApplicationLayout).where(
            ApplicationLayout.status == "active",
            ApplicationLayout.warehouse_bindings.isnot(None)
        )
    )
    for layout in layouts_result.scalars().all():
        try:
            bindings = json.loads(layout.warehouse_bindings) if isinstance(layout.warehouse_bindings, str) else layout.warehouse_bindings
            for ri, bid in bindings.items():
                if bid not in bindings_map:
                    bindings_map[bid] = []
                bindings_map[bid].append(layout.layout_code)
        except Exception:
            pass

    # For deducted items (sheet_count=0), get original quantity from last deduction movement
    deducted_ids = [i.id for i in items if (i.sheet_count or 0) == 0]
    original_qty_map = {}
    if deducted_ids:
        movements_result = await db.execute(
            select(WarehouseMovement).where(
                WarehouseMovement.warehouse_item_id.in_(deducted_ids),
                WarehouseMovement.quantity_change < 0,
            ).order_by(desc(WarehouseMovement.created_at))
        )
        for m in movements_result.scalars().all():
            if m.warehouse_item_id not in original_qty_map:
                original_qty_map[m.warehouse_item_id] = abs(m.quantity_change)

    enriched = []
    for i in items:
        d = _item_to_dict(i)
        d["bound_to"] = bindings_map.get(i.id, [])
        if (i.sheet_count or 0) == 0 and i.id in original_qty_map:
            d["original_sheet_count"] = original_qty_map[i.id]
        enriched.append(d)

    return enriched


@router.get("/balance")
async def get_warehouse_balance(
        metal: str = None,
        grade: str = None,
        sheet_w: float = None,
        sheet_h: float = None,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR, UserRole.ACCOUNTANT))
):
    query = select(WarehouseItem)
    if metal:
        query = query.where(WarehouseItem.metal.ilike(f"%{metal}%"))
    if grade:
        query = query.where(WarehouseItem.grade.ilike(f"%{grade}%"))
    if sheet_w is not None:
        query = query.where(WarehouseItem.sheet_w == sheet_w)
    if sheet_h is not None:
        query = query.where(WarehouseItem.sheet_h == sheet_h)
    result = await db.execute(query.order_by(WarehouseItem.metal, WarehouseItem.grade))
    items = result.scalars().all()
    return [_item_to_dict(i) for i in items]


@router.get("/export")
async def export_warehouse_xlsx(
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR))
):
    from app.services.exporters import export_warehouse

    result = await db.execute(
        select(WarehouseItem).order_by(desc(WarehouseItem.created_at))
    )
    items = result.scalars().all()

    data = [
        {
            "id": i.id,
            "metal": i.metal,
            "grade": i.grade or "",
            "size": i.size or (f"{int(i.sheet_w)}x{int(i.sheet_h)}" if i.sheet_w and i.sheet_h else ""),
            "sheet_count": i.sheet_count,
            "owner": i.owner or "",
            "note": i.note or "",
            "created_at": i.created_at.strftime('%d.%m.%Y') if i.created_at else "",
        }
        for i in items
    ]

    output = export_warehouse(data)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=warehouse.xlsx"}
    )


@router.post("/")
async def create_warehouse_item(
        body: WarehouseItemCreate,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR))
):
    if not body.metal:
        raise HTTPException(status_code=400, detail="Металл обязателен")

    size = body.size
    if not size and body.sheet_w and body.sheet_h:
        size = f"{int(body.sheet_w)}x{int(body.sheet_h)}"

    article = body.article
    if not article:
        base = _generate_article_base(body.metal, body.grade, body.thickness)
        # Find next available number for this base
        existing = await db.execute(
            select(WarehouseItem.article).where(WarehouseItem.article.like(f"{base}/%"))
        )
        used = set()
        for row in existing.scalars():
            try:
                num = int(row.split("/")[-1].split(".")[0])
                used.add(num)
            except (ValueError, IndexError):
                pass
        next_num = 1
        while next_num in used:
            next_num += 1
        article = f"{base}/{next_num}"

    item = WarehouseItem(
        metal=body.metal,
        grade=body.grade,
        thickness=body.thickness,
        size=size,
        sheet_w=body.sheet_w,
        sheet_h=body.sheet_h,
        sheet_count=body.sheet_count,
        weight=body.weight,
        article=article,
        item_type=body.item_type,
        owner=body.owner,
        note=body.note,
        created_by=user.id,
    )

    # Auto-calculate weight from thickness if not provided
    if not item.weight and item.thickness and item.sheet_w and item.sheet_h:
        item.weight = round(item.thickness * item.sheet_w * item.sheet_h * 7.85 / 1_000_000, 2)

    db.add(item)
    await db.flush()

    if body.sheet_count > 0:
        db.add(WarehouseMovement(
            warehouse_item_id=item.id,
            quantity_change=body.sheet_count,
            movement_type="initial",
            reason="Создание записи на складе",
            created_by=user.id,
        ))

    await db.commit()
    await db.refresh(item)

    try:
        from app.main import manager
        await manager.broadcast({
            "type": "notification",
            "message": f"Склад: добавлен {body.metal} (арт. {article})"
        })
    except Exception:
        pass

    return _item_to_dict(item)


@router.patch("/{item_id}")
async def update_warehouse_item(
        item_id: int,
        body: WarehouseItemUpdate,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR))
):
    result = await db.execute(select(WarehouseItem).where(WarehouseItem.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    update_data = body.model_dump(exclude_unset=True)

    if "sheet_w" in update_data or "sheet_h" in update_data:
        new_w = update_data.get("sheet_w", item.sheet_w)
        new_h = update_data.get("sheet_h", item.sheet_h)
        if new_w and new_h and "size" not in update_data:
            update_data["size"] = f"{int(new_w)}x{int(new_h)}"

    for field, value in update_data.items():
        setattr(item, field, value)

    # Auto-recalculate weight from thickness if dimensions or thickness changed
    if not item.weight and item.thickness and item.sheet_w and item.sheet_h:
        item.weight = round(item.thickness * item.sheet_w * item.sheet_h * 7.85 / 1_000_000, 2)

    await db.commit()
    await db.refresh(item)
    return _item_to_dict(item)


@router.post("/{item_id}/deduct")
async def deduct_warehouse_item(
        item_id: int,
        body: WarehouseDeductRequest,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR))
):
    if body.quantity <= 0:
        raise HTTPException(status_code=400, detail="Количество должно быть > 0")

    result = await db.execute(select(WarehouseItem).where(WarehouseItem.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    if item.sheet_count < body.quantity:
        raise HTTPException(status_code=400, detail=f"Недостаточно на складе: {item.sheet_count} листов")

    item.sheet_count -= body.quantity
    item.last_deducted_at = datetime.now(timezone.utc)

    movement = WarehouseMovement(
        warehouse_item_id=item.id,
        application_id=body.application_id,
        quantity_change=-body.quantity,
        movement_type="deduction",
        reason=body.reason,
        created_by=user.id,
    )
    db.add(movement)

    if body.application_id:
        app_result = await db.execute(select(Application).where(Application.id == body.application_id))
        app = app_result.scalar_one_or_none()
        if app:
            app.warehouse_item_id = item.id
            app.sheets_used = body.quantity
            app.warehouse_deducted = True

    # Bind to layout if specified
    if body.layout_id:
        layout_result = await db.execute(select(ApplicationLayout).where(ApplicationLayout.id == body.layout_id))
        layout = layout_result.scalar_one_or_none()
        if layout:
            layout.warehouse_item_id = item.id
            layout.layout_sheets_used = body.quantity

    await db.commit()
    return {"status": "success", "sheet_count": item.sheet_count}


@router.post("/{item_id}/return")
async def return_warehouse_item(
        item_id: int,
        body: WarehouseReturnRequest,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR))
):
    if body.quantity <= 0:
        raise HTTPException(status_code=400, detail="Количество должно быть > 0")

    result = await db.execute(select(WarehouseItem).where(WarehouseItem.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    item.sheet_count += body.quantity

    movement = WarehouseMovement(
        warehouse_item_id=item.id,
        application_id=body.application_id,
        quantity_change=body.quantity,
        movement_type="return",
        reason=body.reason,
        created_by=user.id,
    )
    db.add(movement)

    # Add note with reason, date, and user
    if body.reason:
        from datetime import datetime, timezone
        now = datetime.now(timezone.utc).strftime("%d.%m.%Y %H:%M")
        note = ItemNote(
            item_type="warehouse",
            item_id=item.id,
            user_id=user.id,
            username=user.username,
            text=f"[{now}] Возврат {body.quantity} лист(ов). Причина: {body.reason}",
        )
        db.add(note)

    if body.application_id:
        app_result = await db.execute(select(Application).where(Application.id == body.application_id))
        app = app_result.scalar_one_or_none()
        if app:
            app.warehouse_deducted = False
            app.sheets_used = None
            app.warehouse_item_id = None

    await db.commit()
    return {"status": "success", "sheet_count": item.sheet_count}


@router.get("/{item_id}/movements")
async def get_item_movements(
        item_id: int,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR))
):
    result = await db.execute(
        select(WarehouseItem).where(WarehouseItem.id == item_id)
    )
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    movements_result = await db.execute(
        select(WarehouseMovement)
        .where(WarehouseMovement.warehouse_item_id == item_id)
        .order_by(desc(WarehouseMovement.created_at))
    )
    movements = movements_result.scalars().all()

    creator_names = {}
    for m in movements:
        if m.created_by and m.created_by not in creator_names:
            u_result = await db.execute(select(User.username).where(User.id == m.created_by))
            u_name = u_result.scalar_one_or_none()
            creator_names[m.created_by] = u_name or f"#{m.created_by}"

    return [
        {
            "id": m.id,
            "warehouse_item_id": m.warehouse_item_id,
            "application_id": m.application_id,
            "quantity_change": m.quantity_change,
            "movement_type": m.movement_type,
            "reason": m.reason,
            "created_by": creator_names.get(m.created_by, f"#{m.created_by}" if m.created_by else None),
            "created_at": m.created_at.isoformat() if m.created_at else None,
        }
        for m in movements
    ]


# === Remnants ===

@router.get("/remnants")
async def list_remnants(
        item_id: int = None,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR))
):
    query = select(WarehouseRemnant).where(WarehouseRemnant.is_available == True)
    if item_id:
        query = query.where(WarehouseRemnant.warehouse_item_id == item_id)
    result = await db.execute(query.order_by(desc(WarehouseRemnant.created_at)))
    remnants = result.scalars().all()
    return [
        {
            "id": r.id,
            "warehouse_item_id": r.warehouse_item_id,
            "article": r.article,
            "original_w": r.original_w,
            "original_h": r.original_h,
            "vertices": r.vertices if isinstance(r.vertices, list) else (json.loads(r.vertices) if r.vertices else []),
            "area": r.area,
            "weight": r.weight,
            "is_available": r.is_available,
            "note": r.note,
            "created_at": r.created_at.isoformat() if r.created_at else None,
        }
        for r in remnants
    ]


@router.post("/remnants")
async def create_remnant(
        body: dict,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR))
):
    warehouse_item_id = body.get("warehouse_item_id")
    if not warehouse_item_id:
        raise HTTPException(status_code=400, detail="warehouse_item_id обязателен")

    wh_result = await db.execute(select(WarehouseItem).where(WarehouseItem.id == warehouse_item_id))
    wh_item = wh_result.scalar_one_or_none()
    if not wh_item:
        raise HTTPException(status_code=404, detail="Позиция на складе не найдена")

    original_w = body.get("original_w", wh_item.sheet_w or 0)
    original_h = body.get("original_h", wh_item.sheet_h or 0)
    vertices = body.get("vertices")
    if not vertices:
        # Default: full rectangle
        vertices = [[0, 0], [original_w, 0], [original_w, original_h], [0, original_h]]

    area = _polygon_area(vertices)
    weight = body.get("weight")

    # Generate article
    article = body.get("article")
    if not article:
        base = _generate_article(wh_item.metal, wh_item.grade, original_w, original_h, "remnant")
        article = base
        existing = await db.execute(select(WarehouseRemnant).where(WarehouseRemnant.article == article))
        if existing.scalar_one_or_none():
            for i in range(2, 100):
                article = f"{base}-{i:02d}"
                existing = await db.execute(select(WarehouseRemnant).where(WarehouseRemnant.article == article))
                if not existing.scalar_one_or_none():
                    break

    remnant = WarehouseRemnant(
        warehouse_item_id=warehouse_item_id,
        article=article,
        original_w=original_w,
        original_h=original_h,
        vertices=vertices,
        area=area,
        weight=weight,
        note=body.get("note"),
        created_by=user.id,
    )
    db.add(remnant)
    await db.commit()
    await db.refresh(remnant)

    return {
        "id": remnant.id,
        "article": remnant.article,
        "vertices": vertices,
        "area": area,
    }


@router.post("/remnants/{remnant_id}/split")
async def split_remnant(
        remnant_id: int,
        body: RemnantSplitRequest,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR))
):
    remnant = None
    wh_item = None

    # If warehouse_item_id provided, find or create remnant from full sheet
    if body.warehouse_item_id:
        wh_result = await db.execute(select(WarehouseItem).where(WarehouseItem.id == body.warehouse_item_id))
        wh_item = wh_result.scalar_one_or_none()
        if not wh_item:
            raise HTTPException(status_code=404, detail="Позиция на складе не найдена")

        # Try to find existing remnant
        rem_result = await db.execute(
            select(WarehouseRemnant).where(
                WarehouseRemnant.id == remnant_id,
                WarehouseRemnant.is_available == True
            )
        )
        remnant = rem_result.scalar_one_or_none()

        # If no remnant, create remnant from item's actual shape
        if not remnant:
            if wh_item.vertices and isinstance(wh_item.vertices, list) and len(wh_item.vertices) >= 3:
                full_vertices = wh_item.vertices
            else:
                full_vertices = [[0, 0], [wh_item.sheet_w, 0], [wh_item.sheet_w, wh_item.sheet_h], [0, wh_item.sheet_h]]
            remnant = WarehouseRemnant(
                warehouse_item_id=wh_item.id,
                article=wh_item.article,
                original_w=wh_item.sheet_w,
                original_h=wh_item.sheet_h,
                vertices=full_vertices,
                area=wh_item.sheet_w * wh_item.sheet_h,
                weight=wh_item.weight,
                created_by=user.id,
            )
            db.add(remnant)
            await db.flush()
    else:
        # Original flow: find remnant by ID
        rem_result = await db.execute(
            select(WarehouseRemnant).where(
                WarehouseRemnant.id == remnant_id,
                WarehouseRemnant.is_available == True
            )
        )
        remnant = rem_result.scalar_one_or_none()
        if not remnant:
            raise HTTPException(status_code=404, detail="Остаток не найден или уже использован")

        wh_result = await db.execute(select(WarehouseItem).where(WarehouseItem.id == remnant.warehouse_item_id))
        wh_item = wh_result.scalar_one_or_none()

    vertices = remnant.vertices if isinstance(remnant.vertices, list) else json.loads(remnant.vertices)

    # Subtract cut rectangle from the actual polygon shape
    remaining = _subtract_rect(vertices, body.x, body.y, body.w, body.h)

    if not remaining or _polygon_area(remaining) <= 0:
        raise HTTPException(status_code=400, detail="Выделенная область полностью покрывает остаток")

    # Calculate weight for cut piece
    if remnant.weight and remnant.original_w and remnant.original_h:
        remnant_area = remnant.original_w * remnant.original_h
        cut_weight = remnant.weight * (body.w * body.h / remnant_area) if remnant_area > 0 else None
    elif wh_item and wh_item.thickness:
        cut_weight = round(wh_item.thickness * body.w * body.h * 7.85 / 1_000_000, 2)
    else:
        cut_weight = body.w * body.h * 0.000001 * 7.85

    # Calculate weight for remaining piece
    remaining_area = _polygon_area(remaining)
    if remnant.weight and remnant.original_w and remnant.original_h:
        remnant_total_area = remnant.original_w * remnant.original_h
        remain_weight = remnant.weight * (remaining_area / remnant_total_area) if remnant_total_area > 0 else None
    elif wh_item and wh_item.thickness:
        remain_weight = round(wh_item.thickness * remaining_area / 1_000_000 * 7.85, 2)
    else:
        remain_weight = remaining_area * 0.000001 * 7.85

    # Generate articles for both pieces (children of parent)
    parent_article = wh_item.article if wh_item else None
    if parent_article:
        # Find next available sub-suffixes: parent.1, parent.2, ...
        used_subs = set()
        existing = await db.execute(
            select(WarehouseItem.article).where(WarehouseItem.article.like(f"{parent_article}.%"))
        )
        for row in existing.scalars():
            try:
                sub = int(row.split(".")[-1])
                used_subs.add(sub)
            except (ValueError, IndexError):
                pass

        sub1 = 1
        while sub1 in used_subs:
            sub1 += 1
        sub2 = sub1 + 1
        while sub2 in used_subs:
            sub2 += 1

        cut_article = f"{parent_article}.{sub1}"
        remain_article = f"{parent_article}.{sub2}"
    else:
        base = _generate_article_base(
            wh_item.metal if wh_item else "Сталь", wh_item.grade, wh_item.thickness
        )
        existing = await db.execute(
            select(WarehouseItem.article).where(WarehouseItem.article.like(f"{base}/%"))
        )
        used = set()
        for row in existing.scalars():
            try:
                num = int(row.split("/")[-1].split(".")[0])
                used.add(num)
            except (ValueError, IndexError):
                pass
        next_num = 1
        while next_num in used:
            next_num += 1
        cut_article = f"{base}/{next_num}"
        remain_article = f"{base}/{next_num + 1}"

    # Create cut piece as new warehouse item
    cut_item = WarehouseItem(
        metal=wh_item.metal if wh_item else "Сталь",
        grade=wh_item.grade,
        thickness=wh_item.thickness if wh_item else None,
        sheet_w=body.w,
        sheet_h=body.h,
        size=f"{int(body.w)}x{int(body.h)}",
        sheet_count=1,
        weight=cut_weight,
        article=cut_article,
        parent_article=parent_article,
        parent_sheet_w=wh_item.sheet_w if wh_item else None,
        parent_sheet_h=wh_item.sheet_h if wh_item else None,
        item_type="standard",
        owner=wh_item.owner if wh_item else None,
        created_by=user.id,
    )
    db.add(cut_item)

    # Create remaining piece as new warehouse item (ALWAYS)
    remain_item = None
    rw, rh = rectBounds(remaining) if remaining else (0, 0)
    if rw >= 10 and rh >= 10:
        remain_item = WarehouseItem(
            metal=wh_item.metal if wh_item else "Сталь",
            grade=wh_item.grade,
            thickness=wh_item.thickness if wh_item else None,
            sheet_w=rw,
            sheet_h=rh,
            size=f"{int(rw)}x{int(rh)}",
            sheet_count=1,
            weight=remain_weight,
            article=remain_article,
            parent_article=parent_article,
            parent_sheet_w=wh_item.sheet_w if wh_item else None,
            parent_sheet_h=wh_item.sheet_h if wh_item else None,
            is_rectangular=isRectangle(remaining),
            vertices=remaining,
            item_type="standard",
            owner=wh_item.owner if wh_item else None,
            created_by=user.id,
        )
        db.add(remain_item)

    # Decrement original sheet count by 1 (cut one sheet from the stack)
    if wh_item:
        wh_item.sheet_count = max(0, wh_item.sheet_count - 1)
        wh_item.note = (wh_item.note or "") + f"\nРазрезан: {int(body.w)}x{int(body.h)}"
        # If no sheets left, delete the item and its movements
        if wh_item.sheet_count <= 0:
            movements = await db.execute(
                select(WarehouseMovement).where(WarehouseMovement.warehouse_item_id == wh_item.id)
            )
            for m in movements.scalars().all():
                await db.delete(m)
            await db.flush()
            await db.delete(wh_item)

    # Delete the temporary remnant
    if remnant:
        await db.delete(remnant)

    await db.commit()

    # Refresh to get IDs
    await db.refresh(cut_item)
    if remain_item:
        await db.refresh(remain_item)

    result = {
        "status": "success",
        "cut_item": _item_to_dict(cut_item),
        "remain_item": _item_to_dict(remain_item) if remain_item else None,
    }

    return result


@router.post("/merge-cut")
async def merge_cut_pieces(
        body: dict,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR))
):
    """Reverse a cut: merge two child pieces back into the original parent sheet."""
    item_id_1 = body.get("item_id_1")
    item_id_2 = body.get("item_id_2")

    if not item_id_1 or not item_id_2:
        raise HTTPException(status_code=400, detail="Укажите две позиции для слияния")
    if item_id_1 == item_id_2:
        raise HTTPException(status_code=400, detail="Нужны две разные позиции")

    # Find both items
    result1 = await db.execute(select(WarehouseItem).where(WarehouseItem.id == item_id_1))
    item1 = result1.scalar_one_or_none()
    result2 = await db.execute(select(WarehouseItem).where(WarehouseItem.id == item_id_2))
    item2 = result2.scalar_one_or_none()

    if not item1 or not item2:
        raise HTTPException(status_code=404, detail="Одна или обе позиции не найдены")

    # Both must be children of the same parent
    if not item1.parent_article or not item2.parent_article:
        raise HTTPException(status_code=400, detail="Обе позиции должны быть разрезанными кусками (иметь parent_article)")
    if item1.parent_article != item2.parent_article:
        raise HTTPException(status_code=400, detail=f"Позиции из разных родителей: {item1.parent_article} и {item2.parent_article}")

    # Both must have stock
    if (item1.sheet_count or 0) < 1:
        raise HTTPException(status_code=400, detail=f"Позиция {item1.article} не может быть слияна: списана")
    if (item2.sheet_count or 0) < 1:
        raise HTTPException(status_code=400, detail=f"Позиция {item2.article} не может быть слияна: списана")

    # Check neither is bound to any layout
    layouts_result = await db.execute(
        select(ApplicationLayout).where(
            ApplicationLayout.status == "active",
            ApplicationLayout.warehouse_bindings.isnot(None)
        )
    )
    for layout in layouts_result.scalars().all():
        try:
            bindings = json.loads(layout.warehouse_bindings) if isinstance(layout.warehouse_bindings, str) else layout.warehouse_bindings
            for ri, bid in bindings.items():
                if bid == item1.id:
                    raise HTTPException(status_code=400, detail=f"Позиция {item1.article} закреплена за раскладкой {layout.layout_code}")
                if bid == item2.id:
                    raise HTTPException(status_code=400, detail=f"Позиция {item2.article} закреплена за раскладкой {layout.layout_code}")
        except HTTPException:
            raise
        except Exception:
            pass

    # Get original parent dimensions from children
    parent_w = item1.parent_sheet_w or item2.parent_sheet_w
    parent_h = item1.parent_sheet_h or item2.parent_sheet_h
    if not parent_w or not parent_h:
        raise HTTPException(status_code=400, detail="Не удалось определить размер исходного листа (parent_sheet_w/h отсутствует)")

    # Create parent item
    parent_article = item1.parent_article
    weight_sum = (item1.weight or 0) + (item2.weight or 0)

    # Delete all movements for children and any existing empty item with same article FIRST
    for child in [item1, item2]:
        movements = await db.execute(
            select(WarehouseMovement).where(WarehouseMovement.warehouse_item_id == child.id)
        )
        for m in movements.scalars().all():
            await db.delete(m)

    # Check if an item with the same article already exists
    existing_result = await db.execute(
        select(WarehouseItem).where(WarehouseItem.article == parent_article)
    )
    existing = existing_result.scalar_one_or_none()
    if existing:
        if (existing.sheet_count or 0) > 0:
            raise HTTPException(status_code=400, detail=f"Позиция {parent_article} уже существует на складе ({existing.sheet_count} листов)")
        # Delete empty existing item's movements
        movements = await db.execute(
            select(WarehouseMovement).where(WarehouseMovement.warehouse_item_id == existing.id)
        )
        for m in movements.scalars().all():
            await db.delete(m)
        await db.delete(existing)

    # Delete both children
    for child in [item1, item2]:
        await db.delete(child)

    await db.flush()

    # NOW create parent (after all deletes are flushed)
    parent = WarehouseItem(
        metal=item1.metal,
        grade=item1.grade,
        thickness=item1.thickness,
        sheet_w=parent_w,
        sheet_h=parent_h,
        size=f"{int(parent_w)}x{int(parent_h)}",
        sheet_count=1,
        weight=weight_sum if weight_sum > 0 else None,
        article=parent_article,
        parent_article=None,
        parent_sheet_w=None,
        parent_sheet_h=None,
        item_type="standard",
        owner=item1.owner,
        created_by=user.id,
    )
    db.add(parent)

    await db.commit()
    await db.refresh(parent)

    return {"status": "success", "parent": _item_to_dict(parent)}


@router.delete("/remnants/{remnant_id}")
async def delete_remnant(
        remnant_id: int,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR))
):
    result = await db.execute(select(WarehouseRemnant).where(WarehouseRemnant.id == remnant_id))
    remnant = result.scalar_one_or_none()
    if not remnant:
        raise HTTPException(status_code=404, detail="Остаток не найден")

    await db.delete(remnant)
    await db.commit()
    return {"status": "success"}


def _polygon_area(vertices: list) -> float:
    n = len(vertices)
    if n < 3:
        return 0
    area = 0
    for i in range(n):
        j = (i + 1) % n
        area += vertices[i][0] * vertices[j][1]
        area -= vertices[j][0] * vertices[i][1]
    return abs(area) / 2


def isRectangle(vertices: list) -> bool:
    if not vertices or len(vertices) < 3:
        return False
    area = _polygon_area(vertices)
    if area <= 0:
        return False
    xs = [v[0] for v in vertices]
    ys = [v[1] for v in vertices]
    bbox_area = (max(xs) - min(xs)) * (max(ys) - min(ys))
    return abs(area - bbox_area) < 1.0


def rectBounds(vertices: list) -> tuple:
    xs = [v[0] for v in vertices]
    ys = [v[1] for v in vertices]
    x, y = min(xs), min(ys)
    w, h = max(xs) - x, max(ys) - y
    return w, h


def _is_rect_polygon(vertices: list) -> bool:
    """Check if polygon is a rectangle (axis-aligned, 4 vertices)."""
    if not vertices or len(vertices) != 4:
        return False
    xs = sorted(set(round(v[0], 1) for v in vertices))
    ys = sorted(set(round(v[1], 1) for v in vertices))
    return len(xs) == 2 and len(ys) == 2


def _subtract_rect(polygon: list, rx: float, ry: float, rw: float, rh: float) -> list:
    """Subtract axis-aligned rectangle from polygon, return remaining vertices."""
    if not polygon or len(polygon) < 3:
        return polygon

    # Fast path: rectangular sheet
    if _is_rect_polygon(polygon):
        W = max(v[0] for v in polygon)
        H = max(v[1] for v in polygon)
        x1, y1 = max(0.0, rx), max(0.0, ry)
        x2, y2 = min(W, rx + rw), min(H, ry + rh)
        if x1 >= x2 or y1 >= y2:
            return polygon
        if x1 == 0 and y1 == 0 and x2 == W and y2 == H:
            return []
        tl, tr = x1 == 0, x2 == W
        tt, tb = y1 == 0, y2 == H
        if tl and tt: return [[x2, 0], [W, 0], [W, H], [0, H], [0, y2], [x2, y2]]
        if tr and tt: return [[0, 0], [x1, 0], [x1, y2], [W, y2], [W, H], [0, H]]
        if tl and tb: return [[0, 0], [W, 0], [W, H], [x2, H], [x2, y1], [0, y1]]
        if tr and tb: return [[0, 0], [W, 0], [W, y1], [x1, y1], [x1, H], [0, H]]
        if tt: return [[0, 0], [x1, 0], [x1, y2], [x2, y2], [x2, 0], [W, 0], [W, H], [0, H]]
        if tb: return [[0, 0], [W, 0], [W, H], [x2, H], [x2, y1], [x1, y1], [x1, H], [0, H]]
        if tl: return [[0, 0], [W, 0], [W, H], [0, H], [0, y2], [x2, y2], [x2, y1], [0, y1]]
        if tr: return [[0, 0], [W, 0], [W, y1], [x1, y1], [x1, y2], [W, y2], [W, H], [0, H]]
        # Middle cut — largest remainder rectangle
        pieces = []
        if y1 > 0: pieces.append((0, 0, W, y1))
        if y2 < H: pieces.append((0, y2, W, H - y2))
        if x1 > 0: pieces.append((0, y1, x1, y2 - y1))
        if x2 < W: pieces.append((x2, y1, W - x2, y2 - y1))
        if not pieces: return polygon
        px, py, pw, ph = max(pieces, key=lambda p: p[2] * p[3])
        return [[px, py], [px + pw, py], [px + pw, py + ph], [px, py + ph]]

    # General polygon: clip against OUTSIDE of each rectangle edge
    rx2, ry2 = rx + rw, ry + rh
    def inside(px, py):
        return rx <= px <= rx2 and ry <= py <= ry2

    def clip_edge(poly, axis, val, keep_above):
        """Clip polygon keeping points where (axis coord - val) * keep_above > 0"""
        if not poly:
            return []
        out = []
        n = len(poly)
        for i in range(n):
            curr = poly[i]
            prev = poly[(i - 1) % n]
            cv = curr[0] if axis == 'x' else curr[1]
            pv = prev[0] if axis == 'x' else prev[1]
            c_in = (cv - val) * keep_above > 0
            p_in = (pv - val) * keep_above > 0
            if c_in:
                if not p_in:
                    t = (val - pv) / (cv - pv) if cv != pv else 0
                    out.append([prev[0] + t * (curr[0] - prev[0]), prev[1] + t * (curr[1] - prev[1])])
                out.append(curr)
            elif p_in:
                t = (val - pv) / (cv - pv) if cv != pv else 0
                out.append([prev[0] + t * (curr[0] - prev[0]), prev[1] + t * (curr[1] - prev[1])])
        return out

    # Clip against outside of each edge (union approach via sequential clipping)
    # First: keep only parts above cut (y < ry)
    r = clip_edge(polygon, 'y', ry, -1)
    # Then: keep only parts below cut (y > ry2)
    r2 = clip_edge(polygon, 'y', ry2, 1)
    # Then: keep only parts left of cut (x < rx)
    r3 = clip_edge(polygon, 'x', rx, -1)
    # Then: keep only parts right of cut (x > rx2)
    r4 = clip_edge(polygon, 'x', rx2, 1)
    # Merge: union of all 4 clippings
    all_parts = [p for p in [r, r2, r3, r4] if p and len(p) >= 3]
    if not all_parts:
        return polygon
    # Return the largest part by area
    best = max(all_parts, key=lambda p: abs(_polygon_area(p)))
    return best


def _clip_polygon(subject: list, clip: list) -> list:
    output = subject[:]
    for i in range(len(clip)):
        if not output:
            return []
        input_list = output
        output = []
        edge_start = clip[i]
        edge_end = clip[(i + 1) % len(clip)]
        for j in range(len(input_list)):
            current = input_list[j]
            previous = input_list[(j - 1) % len(input_list)]
            if _is_inside(current, edge_start, edge_end):
                if not _is_inside(previous, edge_start, edge_end):
                    output.append(_intersection(previous, current, edge_start, edge_end))
                output.append(current)
            elif _is_inside(previous, edge_start, edge_end):
                output.append(_intersection(previous, current, edge_start, edge_end))
    return output


def _is_inside(point, edge_start, edge_end):
    return (edge_end[0] - edge_start[0]) * (point[1] - edge_start[1]) - (edge_end[1] - edge_start[1]) * (point[0] - edge_start[0]) >= 0


def _intersection(p1, p2, p3, p4):
    x1, y1 = p1
    x2, y2 = p2
    x3, y3 = p3
    x4, y4 = p4
    denom = (x1 - x2) * (y3 - y4) - (y1 - y2) * (x3 - x4)
    if abs(denom) < 1e-10:
        return p1
    t = ((x1 - x3) * (y3 - y4) - (y1 - y3) * (x3 - x4)) / denom
    return [x1 + t * (x2 - x1), y1 + t * (y2 - y1)]


# === Item Notes (chat-like) ===

@router.get("/notes/{item_type}/{item_id}")
async def get_item_notes(
        item_type: str,
        item_id: int,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR))
):
    if item_type not in ('warehouse', 'deficit'):
        raise HTTPException(status_code=400, detail="Invalid item type")

    result = await db.execute(
        select(ItemNote)
        .where(ItemNote.item_type == item_type, ItemNote.item_id == item_id)
        .order_by(desc(ItemNote.created_at))
    )
    notes = result.scalars().all()
    return [
        {
            "id": n.id,
            "user_id": n.user_id,
            "username": n.username,
            "text": n.text,
            "created_at": n.created_at.isoformat() if n.created_at else None,
        }
        for n in notes
    ]


@router.post("/notes/{item_type}/{item_id}")
async def add_item_note(
        item_type: str,
        item_id: int,
        body: dict,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR))
):
    if item_type not in ('warehouse', 'deficit'):
        raise HTTPException(status_code=400, detail="Invalid item type")

    text = body.get("text", "").strip()
    if not text:
        raise HTTPException(status_code=400, detail="Text is required")

    note = ItemNote(
        item_type=item_type,
        item_id=item_id,
        user_id=user.id,
        username=user.username,
        text=text,
    )
    db.add(note)
    await db.commit()
    await db.refresh(note)

    return {
        "id": note.id,
        "user_id": note.user_id,
        "username": note.username,
        "text": note.text,
        "created_at": note.created_at.isoformat() if note.created_at else None,
    }


@router.delete("/notes/{item_type}/{item_id}/{note_id}")
async def delete_item_note(
        item_type: str,
        item_id: int,
        note_id: int,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN))
):
    result = await db.execute(
        select(ItemNote).where(ItemNote.id == note_id, ItemNote.item_type == item_type, ItemNote.item_id == item_id)
    )
    note = result.scalar_one_or_none()
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")

    await db.delete(note)
    await db.commit()
    return {"status": "success"}


@router.delete("/{item_id}")
async def delete_warehouse_item(
        item_id: int,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR))
):
    result = await db.execute(select(WarehouseItem).where(WarehouseItem.id == item_id))
    item = result.scalar_one_or_none()
    if not item:
        raise HTTPException(status_code=404, detail="Запись не найдена")

    # Delete associated movements first (FK constraint)
    movements = await db.execute(
        select(WarehouseMovement).where(WarehouseMovement.warehouse_item_id == item_id)
    )
    for m in movements.scalars().all():
        await db.delete(m)
    await db.flush()

    await db.delete(item)
    await db.commit()

    try:
        from app.main import manager
        await manager.broadcast({
            "type": "notification",
            "message": f"Склад: удалена запись #{item_id}"
        })
    except Exception:
        pass

    return {"status": "success"}


@router.get("/deficit-analysis")
async def deficit_analysis(
        standard_w: float = 1500,
        standard_h: float = 6000,
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR, UserRole.ACCOUNTANT))
):
    """Анализ дефицита: потребность заказов vs наличие на складе"""
    standard_area = standard_w * standard_h

    # 1. Get all active applications (not cut, not rejected) with layouts
    apps_result = await db.execute(
        select(Application, Customer)
        .join(Customer, Application.customer_id == Customer.id, isouter=True)
        .where(Application.status.notin_(["cut", "rejected"]))
    )
    apps = apps_result.all()

    # 2. For each app, get layouts
    demand = {}  # key: (grade, thickness) -> {total_area, by_customer: {name: area}}
    for app, customer in apps:
        grade = app.steel_grade or ""
        thickness = app.thickness or 0
        key = (grade, thickness)
        cust_name = customer.name if customer else "—"

        layouts_result = await db.execute(
            select(ApplicationLayout).where(
                ApplicationLayout.application_id == app.id,
                ApplicationLayout.status == "active"
            )
        )
        layouts = layouts_result.scalars().all()

        if key not in demand:
            demand[key] = {"total_area": 0, "total_sheets": 0, "by_customer": {}}
        if cust_name not in demand[key]["by_customer"]:
            demand[key]["by_customer"][cust_name] = {"area": 0, "sheets": 0}

        for layout in layouts:
            area = (layout.sheet_w or 0) * (layout.sheet_h or 0) * (layout.sheet_count or 1)
            demand[key]["total_area"] += area
            demand[key]["by_customer"][cust_name]["area"] += area
            demand[key]["total_sheets"] += layout.sheet_count or 1
            demand[key]["by_customer"][cust_name]["sheets"] += layout.sheet_count or 1

    # 3. Warehouse stock grouped by (grade, thickness)
    stock_result = await db.execute(
        select(WarehouseItem).where(WarehouseItem.sheet_count > 0)
    )
    stock_items = stock_result.scalars().all()

    stock = {}  # key: (grade, thickness) -> {total_sheets, total_area, by_customer: {name: sheets, articles}}
    for item in stock_items:
        grade = item.grade or ""
        thickness = item.thickness or 0
        key = (grade, thickness)
        cust_name = item.owner or "—"
        sheets = item.sheet_count or 0
        area = (item.sheet_w or 0) * (item.sheet_h or 0) * sheets

        if key not in stock:
            stock[key] = {"total_sheets": 0, "total_area": 0, "by_customer": {}}
        if cust_name not in stock[key]["by_customer"]:
            stock[key]["by_customer"][cust_name] = {"sheets": 0, "area": 0, "articles": []}

        stock[key]["total_sheets"] += sheets
        stock[key]["total_area"] += area
        stock[key]["by_customer"][cust_name]["sheets"] += sheets
        stock[key]["by_customer"][cust_name]["area"] += area
        if item.article:
            stock[key]["by_customer"][cust_name]["articles"].append(item.article)

    # 4. Build deficit table
    all_keys = sorted(set(list(demand.keys()) + list(stock.keys())), key=lambda k: (k[0], k[1] or 0))

    deficit = []
    for key in all_keys:
        grade, thickness = key
        d = demand.get(key, {"total_area": 0, "total_sheets": 0, "by_customer": {}})
        s = stock.get(key, {"total_area": 0, "total_sheets": 0, "by_customer": {}})

        # Convert areas to standard sheets
        d_sheets_std = d["total_area"] / standard_area if standard_area > 0 else 0
        s_sheets_std = s["total_area"] / standard_area if standard_area > 0 else 0

        deficit_sheets = s["total_sheets"] - d_sheets_std

        deficit.append({
            "grade": grade,
            "thickness": thickness,
            "demand_area": d["total_area"],
            "demand_sheets_std": round(d_sheets_std, 1),
            "demand_by_customer": {k: {"area": v["area"], "sheets_std": round(v["area"] / standard_area, 1) if standard_area > 0 else 0} for k, v in d["by_customer"].items()},
            "stock_sheets": s["total_sheets"],
            "stock_area": s["total_area"],
            "stock_by_customer": {k: {"sheets": v["sheets"], "area": v["area"], "articles": v.get("articles", [])} for k, v in s["by_customer"].items()},
            "deficit_sheets": round(deficit_sheets, 1),
        })

    return {
        "standard_w": standard_w,
        "standard_h": standard_h,
        "deficit": deficit,
    }


@router.get("/deficit-export")
async def deficit_export(
        standard_w: float = 1500,
        standard_h: float = 6000,
        grade: str = None,
        thickness: str = None,
        customer_filter: str = Query(None, alias="customer"),
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.DIRECTOR, UserRole.OPERATOR, UserRole.ACCOUNTANT))
):
    """Экспорт дефицита в Excel"""
    from app.services.exporters import style_header
    from openpyxl import Workbook
    from io import BytesIO

    # Reuse deficit logic
    standard_area = standard_w * standard_h

    apps_result = await db.execute(
        select(Application, Customer)
        .join(Customer, Application.customer_id == Customer.id, isouter=True)
        .where(Application.status.notin_(["cut", "rejected"]))
    )
    apps = apps_result.all()

    demand = {}
    for app, customer in apps:
        grade_val = app.steel_grade or ""
        thickness_val = app.thickness or 0
        key = (grade_val, thickness_val)
        cust_name = customer.name if customer else "—"

        layouts_result = await db.execute(
            select(ApplicationLayout).where(
                ApplicationLayout.application_id == app.id,
                ApplicationLayout.status == "active"
            )
        )
        layouts = layouts_result.scalars().all()

        if key not in demand:
            demand[key] = {"total_area": 0, "total_sheets": 0, "by_customer": {}}
        if cust_name not in demand[key]["by_customer"]:
            demand[key]["by_customer"][cust_name] = {"area": 0, "sheets": 0}

        for layout in layouts:
            area = (layout.sheet_w or 0) * (layout.sheet_h or 0) * (layout.sheet_count or 1)
            demand[key]["total_area"] += area
            demand[key]["by_customer"][cust_name]["area"] += area

    stock_result = await db.execute(select(WarehouseItem).where(WarehouseItem.sheet_count > 0))
    stock_items = stock_result.scalars().all()

    stock = {}
    for item in stock_items:
        grade_val = item.grade or ""
        thickness_val = item.thickness or 0
        key = (grade_val, thickness_val)
        cust_name = item.owner or "—"
        sheets = item.sheet_count or 0
        area = (item.sheet_w or 0) * (item.sheet_h or 0) * sheets

        if key not in stock:
            stock[key] = {"total_sheets": 0, "total_area": 0, "by_customer": {}}
        if cust_name not in stock[key]["by_customer"]:
            stock[key]["by_customer"][cust_name] = {"sheets": 0, "area": 0}
        stock[key]["total_sheets"] += sheets
        stock[key]["total_area"] += area
        stock[key]["by_customer"][cust_name]["sheets"] += sheets
        stock[key]["by_customer"][cust_name]["area"] += area

    all_keys = sorted(set(list(demand.keys()) + list(stock.keys())), key=lambda k: (k[0], k[1] or 0))

    # Filter
    filter_grades = [g.strip() for g in grade.split(",")] if grade else []
    filter_thicknesses = [t.strip() for t in thickness.split(",")] if thickness else []
    filter_customers = [c.strip() for c in customer_filter.split(",")] if customer_filter else []

    rows = []
    for key in all_keys:
        g, t = key
        if filter_grades and (g or '—') not in filter_grades:
            continue
        if filter_thicknesses and (f"{t}мм" if t else '—') not in filter_thicknesses:
            continue
        if filter_customers:
            d_custs = set(demand.get(key, {}).get("by_customer", {}).keys())
            s_custs = set(stock.get(key, {}).get("by_customer", {}).keys())
            if not any(c in d_custs or c in s_custs for c in filter_customers):
                continue

        d = demand.get(key, {"total_area": 0, "by_customer": {}})
        s = stock.get(key, {"total_area": 0, "total_sheets": 0, "by_customer": {}})
        d_sheets = d["total_area"] / standard_area if standard_area > 0 else 0
        deficit_val = s["total_sheets"] - d_sheets

        rows.append({
            "grade": g or "—",
            "thickness": f"{t}мм" if t else "—",
            "demand_sheets": round(d_sheets, 1),
            "demand_area": round(d["total_area"] / 1000000, 2),
            "demand_by_customer": d["by_customer"],
            "stock_sheets": s["total_sheets"],
            "stock_area": round(s["total_area"] / 1000000, 2),
            "stock_by_customer": s["by_customer"],
            "balance": round(deficit_val, 1),
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Дефицит"

    headers = ["Марка", "Толщ.", f"Заказы ({standard_w}x{standard_h})", "Заказы м²", "Склад листов", "Склад м²", "Баланс", "Тип", "Клиент", "Листов", "м²", "Артикулы"]
    style_header(ws, headers)

    row_idx = 2
    for row in rows:
        # Main summary row
        ws.cell(row=row_idx, column=1, value=row["grade"])
        ws.cell(row=row_idx, column=2, value=row["thickness"])
        ws.cell(row=row_idx, column=3, value=row["demand_sheets"])
        ws.cell(row=row_idx, column=4, value=row["demand_area"])
        ws.cell(row=row_idx, column=5, value=row["stock_sheets"])
        ws.cell(row=row_idx, column=6, value=row["stock_area"])
        ws.cell(row=row_idx, column=7, value=row["balance"])
        ws.cell(row=row_idx, column=8, value="ИТОГО")
        # Bold summary row
        from openpyxl.styles import Font as XlFont
        for c in range(1, 12):
            ws.cell(row=row_idx, column=c).font = XlFont(bold=True)
        row_idx += 1

        # Demand detail rows
        for cust_name, cust_data in sorted((row["demand_by_customer"] or {}).items()):
            ws.cell(row=row_idx, column=8, value="Заказ")
            ws.cell(row=row_idx, column=9, value=cust_name)
            ws.cell(row=row_idx, column=10, value=round(cust_data.get("sheets_std", 0), 1))
            ws.cell(row=row_idx, column=11, value=round(cust_data.get("area", 0) / 1000000, 2))
            row_idx += 1

        # Stock detail rows
        for cust_name, cust_data in sorted((row["stock_by_customer"] or {}).items()):
            ws.cell(row=row_idx, column=8, value="Склад")
            ws.cell(row=row_idx, column=9, value=cust_name)
            ws.cell(row=row_idx, column=10, value=cust_data.get("sheets", 0))
            ws.cell(row=row_idx, column=11, value=round(cust_data.get("area", 0) / 1000000, 2))
            articles = cust_data.get("articles", [])
            if articles:
                ws.cell(row=row_idx, column=12, value=", ".join(articles))
            row_idx += 1

        # Separator row
        row_idx += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=deficit.xlsx"}
    )


@router.get("/reserved")
async def get_reserved_items(
        db: AsyncSession = Depends(get_db),
        user: User = Depends(require_role(UserRole.ADMIN, UserRole.OPERATOR))
):
    """Return map of warehouse_item_id → list of layout codes that reserve it."""
    layouts_result = await db.execute(
        select(ApplicationLayout).where(
            ApplicationLayout.status == "active",
            ApplicationLayout.warehouse_bindings.isnot(None)
        )
    )
    reserved = {}
    for layout in layouts_result.scalars().all():
        try:
            bindings = json.loads(layout.warehouse_bindings) if isinstance(layout.warehouse_bindings, str) else layout.warehouse_bindings
            runs = json.loads(layout.completed_runs) if layout.completed_runs else []
            for ri, bid in bindings.items():
                ri_int = int(ri)
                # Reserved if run is not yet cut
                if ri_int >= len(runs) or not runs[ri_int]:
                    if bid not in reserved:
                        reserved[bid] = []
                    reserved[bid].append({
                        "layout_code": layout.layout_code,
                        "run_index": ri_int
                    })
        except Exception:
            pass
    return reserved
