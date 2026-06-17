from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO


def style_header(ws, headers, col_widths=None):
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    if col_widths:
        for col_idx, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else 'A'].width = width


def export_applications(applications):
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["ID", "Заказчик", "Заявка", "Материал", "Марка", "Толщ.", "Дав.мат", "Станок", "Кол-во дет.", "Вес", "Статус", "Приоритет", "Дата"]
    style_header(ws, headers)

    for row_idx, app in enumerate(applications, 2):
        ws.cell(row=row_idx, column=1, value=app.get("id"))
        ws.cell(row=row_idx, column=2, value=app.get("customer", ""))
        ws.cell(row=row_idx, column=3, value=app.get("order_name", ""))
        ws.cell(row=row_idx, column=4, value=app.get("material", ""))
        ws.cell(row=row_idx, column=5, value=app.get("steel_grade", ""))
        ws.cell(row=row_idx, column=6, value=app.get("thickness"))
        ws.cell(row=row_idx, column=7, value="Да" if app.get("supply_material") else "Нет" if app.get("supply_material") is False else "-")
        ws.cell(row=row_idx, column=8, value=app.get("machine", ""))
        ws.cell(row=row_idx, column=9, value=app.get("total_parts"))
        ws.cell(row=row_idx, column=10, value=app.get("total_weight"))
        ws.cell(row=row_idx, column=11, value=app.get("status", ""))
        ws.cell(row=row_idx, column=12, value=app.get("priority", ""))
        ws.cell(row=row_idx, column=13, value=app.get("created_at", ""))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_warehouse(items):
    wb = Workbook()
    ws = wb.active
    ws.title = "Склад"

    headers = ["ID", "Металл", "Марка", "Размер", "Кол-во листов", "Владелец", "Примечание", "Дата"]
    style_header(ws, headers)

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.get("id"))
        ws.cell(row=row_idx, column=2, value=item.get("metal", ""))
        ws.cell(row=row_idx, column=3, value=item.get("grade", ""))
        ws.cell(row=row_idx, column=4, value=item.get("size", ""))
        ws.cell(row=row_idx, column=5, value=item.get("sheet_count", 0))
        ws.cell(row=row_idx, column=6, value=item.get("owner", ""))
        ws.cell(row=row_idx, column=7, value=item.get("note", ""))
        ws.cell(row=row_idx, column=8, value=item.get("created_at", ""))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_deficit(items):
    wb = Workbook()
    ws = wb.active
    ws.title = "Дефицит"

    headers = ["ID", "Материал", "Толщ.", "Размер", "Кол-во", "Заказчик", "Примечание", "Статус", "Дата"]
    style_header(ws, headers)

    for row_idx, item in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=item.get("id"))
        ws.cell(row=row_idx, column=2, value=item.get("material", ""))
        ws.cell(row=row_idx, column=3, value=item.get("thickness"))
        ws.cell(row=row_idx, column=4, value=item.get("size", ""))
        ws.cell(row=row_idx, column=5, value=item.get("quantity"))
        ws.cell(row=row_idx, column=6, value=item.get("customer_name", ""))
        ws.cell(row=row_idx, column=7, value=item.get("note", ""))
        ws.cell(row=row_idx, column=8, value=item.get("status", ""))
        ws.cell(row=row_idx, column=9, value=item.get("created_at", ""))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
